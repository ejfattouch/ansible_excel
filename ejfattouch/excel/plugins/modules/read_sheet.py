#!/usr/bin/python
# Copyright (c) 2026 Edward-Joseph Fattouch (ejfattouch@outlook.com)
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = r'''
module: read_sheet
author:
    - Edward-Joseph Fattouch (@ejfattouch)
short_description: Reads data from a single sheet in an Excel document
description:
    - Reads data from a single sheet in an Excel document using the openpyxl module.
requirements:
    - "openpyxl"
    - "xlwings module and running Excel instance on windows and macOS (only required for data evaluation)"
options:
  path:
    description:
      - Path to the Excel document.
    type: str
  sheet:
    description:
      - The name of the sheet to read from.
      - If left empty, first sheet in the workbook will be used.
    type: str
    default: ''
  evaluate:
    description:
      - Whether or not to evaluate the functions in an Excel document. If false, will return the last calculated value.
      - Only compatible on Windows and MacOS with xlwings and an installed Excel instance
    type: bool
    default: false
'''

EXAMPLES = r"""
- name: Read sheet Sheet1 in an Excel document
  ejfattouch.excel.read_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
  register: document

- name: Read first sheet in an Excel document
  ejfattouch.excel.read_sheet:
    path: /your/path/excel/document.xlsx
  register: document
  
- name: Read evaluated sheet Sheet1 in an Excel document
  ejfattouch.excel.read_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    evaluate: true
  register: document
"""

RETURN = r"""
content:
    description: The contents of the sheet.
    type: list
    sample: ["Content 1", "Content 2", 1234]
    returned: always
evaluated:
    description: Returns True if the functions were evaluated.
    type: bool
    returned: always
path:
    description: The path to the Excel document.
    type: str
    returned: always
sheet:
    description: Name of the sheet that was read.
    type: str
    returned: always
"""

import os
from ansible.module_utils.basic import AnsibleModule
# noinspection PyUnresolvedReferences
from ansible_collections.ejfattouch.excel.plugins.module_utils.excel_common import (
    check_excel_installation,
    evaluate_workbook_formulas,
    read_sheet_data,
)
from openpyxl import load_workbook


def main():
    module_args = dict(path=dict(type='path', required=True),
                       sheet=dict(type='str', required=False, default=''),
                       evaluate=dict(type='bool', default=False))
    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True,
    )

    path = module.params['path']
    if not os.path.isfile(path):
        module.fail_json(msg="The specified excel file does not exist at " + path)

    results = {'path': os.path.abspath(path)}

    # Load workbook and resolve sheet name
    try:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
    except Exception as e:
        module.fail_json(msg=f"Failed to open workbook: {e}")

    sheet_name = module.params['sheet'] or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        wb.close()
        module.fail_json(msg=f"Worksheet '{sheet_name}' does not exist in Excel workbook {results['path']}")

    if module.params['evaluate']:
        wb.close()
        try:
            if not check_excel_installation():
                module.fail_json(msg="Excel is not installed, needed for function evaluation.")
            evaluate_workbook_formulas(path)
            # Reload workbook after evaluation to get computed values
            wb = load_workbook(filename=path, data_only=True, read_only=True)
            results['content'] = read_sheet_data(wb[sheet_name])
            results['evaluated'] = True
        except RuntimeError as e:
            module.fail_json(msg=str(e))
        except ModuleNotFoundError as e:
            module.fail_json(msg=f"{e.name} is not installed, needed for function evaluation.")
    else:
        results['content'] = read_sheet_data(wb[sheet_name])
        results['evaluated'] = False

    wb.close()
    results['sheet'] = sheet_name
    module.exit_json(**results)


if __name__ == '__main__':
    main()
