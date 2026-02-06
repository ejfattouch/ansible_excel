#!/usr/bin/python
# Copyright (c) 2026 Edward-Joseph Fattouch (ejfattouch@outlook.com)
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = r'''
module: write_sheet
author:
    - Edward-Joseph Fattouch (@ejfattouch)
short_description: Writes data to a single sheet in an Excel document
description:
    - Writes data to a sheet in an Excel document using the openpyxl module.
requirements:
    - "openpyxl"
    - "xlwings module and running Excel instance on windows and macOS (only required for data evaluation)"
options:
  path:
    description:
      - Path to the Excel document.
    type: str
  sheet:
    description:
      - The name of the sheet to write to. If the sheet does not exist, it will be created.
      - If left empty, first sheet in the workbook will be used.
    type: str
    default: 'First sheet in the workbook'
  cell:
    description:
      - The cell to start writing at.
    type: str
    default: 'A1'
  data:
    description:
      - The data to be written.
    type: list
    default: ''
  override:
    description:
      - Override current data. If set to false, only empty cells will be written to.
    type: bool
    default: true
  evaluate:
    description:
      - Whether or not to evaluate the functions in an Excel document. If false, will return the last calculated value.
      - Only compatible on Windows and MacOS with xlwings and an installed Excel instance
    type: bool
    default: false
'''

EXAMPLES = r"""
- name: Write data to a single cell in an Excel document
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    cell: B10
    data: "your_data"

- name: Write list data to a row starting at B10
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    cell: B10
    data: ["B10_data", "B11_data", "B12_data", ...]

- name: Write multi-row data from 2d list starting at B10
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    cell: B10
    data: [["B10_data", "B11_data", "B12_data"], ["C10_data", "C11_data", "C12_data"], ...]

- name: Write data to default cell (A1)
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    data: "your_data"

- name: Write data to first sheet
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    data: "your_data"

- name: Write data only on empty cells (preserves already filled cells)
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    cell: B10
    data: "your_data"
    override: false

- name: Write data and re-evaluate
  ejfattouch.excel.write_sheet:
    path: /your/path/excel/document.xlsx
    sheet: "Sheet1"
    cell: B10
    data: "your_data"
    evaluate: true
"""

RETURN = r"""
cells:
    description: List of cells that were changed.
    type: list
    content: str
    returned: always
evaluated:
    description: Returns whether or not the functions were evaluated.
    type: bool
    returned: always
path:
    description: The path to the Excel document.
    type: str
    returned: always
sheet:
    description: Name of the sheet that was written to.
    type: str
    returned: always
"""

import os
from ansible.module_utils.basic import AnsibleModule
# noinspection PyUnresolvedReferences
from ansible_collections.ejfattouch.excel.plugins.module_utils.excel_common import (
    check_excel_installation,
    evaluate_workbook_formulas,
)
from openpyxl import load_workbook
import openpyxl.utils as xl_utils
from openpyxl.cell.cell import MergedCell


# Grabs sheet from Excel workbook or creates a new one if it doesn't already exist
def grab_sheet(sheet_name, wb):
    sheets = wb.sheetnames
    if sheet_name not in sheets:
        wb.create_sheet(sheet_name)
    return wb[sheet_name]


# Performs data_validation tasks before writing to the document to ensure data is consistent
def validate_data(data_list):
    if len(data_list) == 0:
        raise Exception("Empty Data List")

    if isinstance(data_list[0], (str, int, float)):  # Parse through data to check if all values are strings or nums
        if not all(isinstance(element, (str, int, float)) for element in data_list):
            raise TypeError("Data must be of type str or number across the whole list")
        return True
    elif type(data_list[0]) is list:
        for sub_list in data_list:
            if type(sub_list) is not list:
                raise TypeError("All elements in a list of sublists must be of type list")
            if not all(isinstance(element, (str, int, float)) for element in sub_list):
                raise TypeError("Data must be of type str or number across the sublist")
        return True
    else:
        raise TypeError("data must be a list containing only str or number or a 2d list")


def write_data_to_sheet(data, cell, wb, sheet_name, override=False):
    ws = grab_sheet(sheet_name, wb)
    cell_coord = xl_utils.coordinate_to_tuple(cell)
    start_row, start_col = cell_coord[0], cell_coord[1]

    cell_list = []

    def write_cell(value, row, column):  # Returns true if successfully written
        w_cell = ws.cell(row=row, column=column)
        if isinstance(w_cell, MergedCell):  # Skip over merged cells when writing
            return False
        if w_cell.value is None:
            w_cell.value = value
            return True
        elif override and w_cell.value != value:
            w_cell.value = value
            return True
        return "skipped"

    def write_row(row_elements, row, col=start_col):
        num_items = len(row_elements)
        item_number = 0
        write_change = False
        while item_number < num_items:
            status = write_cell(row_elements[item_number], row, col)
            if status == "skipped":  # If writing data was skipped move on to the next item
                item_number += 1
            elif status:
                write_change = True
                item_number += 1
                cell_list.append(xl_utils.get_column_letter(col) + str(row))
            col += 1
        return write_change

    changed = False
    if isinstance(data[0], (str, int, float)):  # write in a single row
        changed = write_row(data, start_row)
    elif type(data[0]) is list:
        for sub_list in data:
            if write_row(sub_list, start_row):
                changed = True
            start_row += 1
    cell_list.sort()
    return changed, cell_list


def main():
    module_args = dict(path=dict(type='path', required=True),
                       sheet=dict(type='str', required=False, default=''),
                       cell=dict(type='str', required=False, default='A1'),
                       data=dict(type='list', required=True),
                       override=dict(type='bool', required=False, default=True),
                       evaluate=dict(type='bool', default=False))

    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True,
    )

    filepath = module.params['path']
    data = module.params['data']
    sheet_name = module.params['sheet']
    cell = module.params['cell']
    override = module.params['override']
    evaluate = module.params['evaluate']

    if not os.path.isfile(filepath):
        module.fail_json(msg="The specified excel file does not exist at " + filepath)

    try:
        validate_data(data)
    except (TypeError, Exception) as e:
        module.fail_json(msg=str(e))

    # Load workbook without data_only to preserve formulas
    try:
        wb = load_workbook(filename=filepath)
    except Exception as e:
        module.fail_json(msg=f"Failed to open workbook: {e}")

    sheet_name = sheet_name or wb.sheetnames[0]

    cell_changed_list = []
    try:
        changed, cell_changed_list = write_data_to_sheet(data, cell, wb, sheet_name, override)
    except ValueError as e:
        wb.close()
        module.fail_json(msg=f"Incorrect value for cell '{cell}': {e}")

    if changed:
        wb.save(filepath)
    wb.close()

    evaluated = False
    if evaluate:
        try:
            if not check_excel_installation():
                module.fail_json(msg="Excel is not installed, needed for function evaluation.")
            evaluate_workbook_formulas(filepath)
            evaluated = True
        except RuntimeError as e:
            module.fail_json(msg=str(e))
        except ModuleNotFoundError as e:
            module.fail_json(msg=f"{e.name} is not installed, needed for function evaluation.")

    results = {
        'path': os.path.abspath(filepath),
        'sheet': sheet_name,
        'cells': cell_changed_list,
        'evaluated': evaluated,
        'changed': changed,
    }

    module.exit_json(**results)


if __name__ == '__main__':
    main()
