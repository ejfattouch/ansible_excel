#!/usr/bin/python
# Copyright (c) 2024 Edward-Joseph Fattouch (ejfattouch@outlook.com)
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = r'''
module: read_document
author:
    - Edward-Joseph Fattouch (@ejfattouch)
short_description: Writes data to a single sheet in an Excel document
description:
    - Writes data to a sheet in an Excel document using the openpyxl module.
requirements:
    - "openpyxl"
    - "xlwings module and running Excel instance on windows and macOS (only required for data evaluation)"
options:
  path:
    description:
      - Path to the Excel document.
    type: str
  sheet:
    description:
      - The name of the sheet to write to. If the sheet does not exist, it will be created.
    notes: If left empty, first sheet in the workbook will be used.
    type: str
    default: ''
  cell:
    description:
      - The cell to start writing at.
    type: str
    default: 'A1'
  data:
    description:
      - The data to be written.
    type: list
    default: ''
  override:
    description:
      - Override current data. If set to false, only empty cells will be written to.
    type: bool
    default: true
  evaluate:
    description:
      - Whether or not to evaluate the functions in an Excel document after all data has been written.
    notes: Only compatible on Windows and MacOS with xlwings and an installed Excel instance
    type: bool
    default: false
'''

EXAMPLES = r"""
- name: Read an entire Excel document
  ejf.read_excel_document:
    path: /your/path/excel/document.xlsx
  register: document
  
- name: Read an Excel document with its values evaluated
  ejf.read_excel_document:
    path: /your/path/excel/document.xlsx
    evaluate: true
  register: document
"""

RETURN = r"""
cells:
    description: List of cells that were changed.
    type: list
    content: str
    returned: always
evaluated:
    description: Returns whether or not the functions were evaluated.
    type: bool
    returned: always
overridden:
    description: Returns whether or not the non-empty cells were overriden.
    type: bool
    returned: always
path:
    description: The path to the Excel document.
    type: str
    returned: always
sheet:
    description: Name of the sheet that was written to.
    type: str
    returned: always
"""

import os
from ansible.module_utils.basic import AnsibleModule
from openpyxl import *
import openpyxl.utils as xl_utils
from openpyxl.cell.cell import MergedCell


def check_excel_installation():
    import platform
    # Checks if os is Windows or darwin (macOS)
    if platform.system() == 'Windows':
        try:
            # Check if Excel executable exists
            excel_path = os.path.join(os.environ["ProgramFiles"], "Microsoft Office", "root", "Office16", "EXCEL.EXE")
            excel_path_x86 = os.path.join(os.environ["ProgramFiles(x86)"], "Microsoft Office", "root", "Office16",
                                          "EXCEL.EXE")
            return os.path.exists(excel_path) or os.path.exists(excel_path_x86)
        except KeyError:
            # Environment variable not found (Excel not installed)
            return False
    elif platform.system() == 'Darwin':
        excel_path = os.path.join("/Applications", "Microsoft Excel.app")
        return os.path.exists(excel_path)
    else:
        raise RuntimeError("Only Windows and MacOS are supported for evaluation")


# Grabs sheet from Excel workbook or creates a new one if it doesn't already exist
def grab_sheet(sheet_name, wb):
    sheets = wb.sheetsnames
    if sheet_name in sheets:
        return wb[sheet_name]
    wb.create_sheet(sheet_name)
    return wb[sheet_name]


# Performs data_validation tasks before writing to the document to ensure data is consistent
def validate_data(data_list):
    if len(data_list) == 0:
        raise Exception("Empty Data List")

    if isinstance(data_list[0], (str, int, float)):  # Parse through data to check if all values are strings or nums
        if not all(isinstance(element, (str, int, float)) for element in data_list):
            raise TypeError("Data must be of type str or number across the whole list")
        return True
    elif type(data_list[0]) is list:
        for sub_list in data_list:
            if type(sub_list) is not list:
                raise TypeError("All elements in a list of sublists must be of type list")
            if not all(isinstance(element, (str, int, float)) for element in sub_list):
                raise TypeError("Data must be of type str or number across the sublist")
        return True
    else:
        raise TypeError("data must be a list containing only str or number or a 2d list")


def write_data_to_sheet(data, cell, wb, sheet_name, override=False):
    ws = wb[sheet_name]
    cell_coord = xl_utils.coordinate_to_tuple(cell)
    start_row, start_col = cell_coord[0], cell_coord[1]

    def write_cell(value, row, column):  # Returns true if successfully written
        w_cell = ws.cell(row=row, column=column)
        if isinstance(w_cell, MergedCell):  # Skip over merged cells when writing
            return False, False
        if cell.value is None:
            cell.value = value
            return True, False
        elif override and cell.value != value:
            cell.value = value
            return True, True
        return "skipped", False

    def write_row(row_elements, row, col=start_col):
        num_items = len(row_elements)
        item_number = 0
        write_change = False
        while item_number < num_items:
            status = write_cell(row_elements[item_number], row, col)
            if status[0] == "skipped":
                item_number += 1
                col += 1
            elif status[0]:
                write_change = True
                item_number += 1
                col += 1
            else:  # If unsuccessful move on to next cell
                col += 1
        return write_change

    changed = False
    if isinstance(data[0], (str, int, float)):  # write in a single row
        changed = write_row(data, start_row)
    elif type(data[0]) is list:
        for sub_list in data:
            if write_row(sub_list, start_row):
                changed = True
            start_row += 1
    return changed


def evaluate_workbook(path):
    import xlwings as xw
    # xlwings allows for workbook to be opened using running excel instance
    excel_app = xw.App(visible=False)
    excel_book = excel_app.books.open(path)
    excel_book.save()  # calling excel save compute functions and store it in cache
    excel_book.close()
    excel_app.quit()


def main():
    module_args = dict(path=dict(type='path', required=True),
                       sheet=dict(type='str', required=False, default=''),
                       cell=dict(type='str', required=False, default='A1'),
                       data=dict(type='list', required=True),
                       override=dict(type='bool', required=False, default=True),
                       evaluate=dict(type='bool', default=False))

    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True,
    )

    filepath = module.params['path']
    data = module.params['data']
    sheet_name = module.params['sheet']
    cell = module.params['cell']
    override = module.params['override']
    evaluate = module.params['evaluate']

    results = {}
    changed = False

    if not os.path.isfile(filepath):
        module.fail_json(msg="The specified excel file does not exist at " + filepath)
        return 1

    try:
        validate_data(data)
    except TypeError as e:  # Inconsistent data type
        module.fail_json(msg=str(e))
        return 1
    except Exception as e:  # Empty data
        module.fail_json(msg=str(e))
        return 1

    wb = load_workbook(filename=filepath, data_only=True)
    if not sheet_name:
        sheet_name = wb.sheetnames[0]  # Get the first sheet if the sheet_name is unspecified

    changed = write_data_to_sheet(data, cell, wb, sheet_name, override)

    evaluated = False
    if evaluate:
        try:
            if not check_excel_installation():
                module.fail_json(msg="Excel is not installed, needed for function evaluation.")
                return 1
            evaluate_workbook(filepath)
            evaluated = True
        except RuntimeError as e:
            module.fail_json(msg=str(e))
            return 1
        except ModuleNotFoundError as e:
            module.fail_json(msg=f"{e.name} is not installed, needed for function evaluation.")
            return 1
        except KeyError as e:
            module.fail_json(msg=f"{e} does not exist at {os.path.abspath(filepath)}")
            return 1

    results['path'] = os.path.abspath(filepath)
    results['sheet'] = sheet_name
    results['cell'] = cell
    results['overridden'] = was_overridden
    results['evaluate'] = evaluated

    module.exit_json(**results)
    return 0


if __name__ == '__main__':
    main()
