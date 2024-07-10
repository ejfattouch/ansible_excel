#!/usr/bin/python
# Copyright (c) 2024 Edward-Joseph Fattouch (ejfattouch@outlook.com)
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = r'''
module: read_document
author:
    - Edward-Joseph Fattouch (@ejfattouch)
short_description: Reads an entire excel document
description:
    - Reads an entire excel document using the openpyxl module.
requirements:
    - "openpyxl"
    - "xlwings"
    - "running excel instance on windows and macOS (only required for data evaluation)"
options:
  path:
    description:
      - Path to the excel document.
    type: str
  evaluate:
    description:
      - Whether or not to evaluate the functions in an excel document.
      notes: Only compatible on Windows and MacOS with xlwings and a running excel instance
      type: bool
      default: false
'''

EXAMPLES = r"""
- name: Read an entire excel document
  ejf.read_excel_document:
    path: /your/path/excel/document.xlsx
  register: document
  
- name: Read an excel document with its values evaluated
  ejf.read_excel_document:
    path: /your/path/excel/document.xlsx
    evaluate: true
  register: document
"""

RETURN = r"""
path:
    description: The path to the excel document.
    type: str
    returned: always
sheets:
    description: List containing the names of the sheets.
    type: list
    returned: always
content:
    description: The contents of each sheets of the document.
    type: dict
    sample: {'Sheet1': [...], 'Sheet2': [...], ...}
"""

import os
from ansible.module_utils.basic import AnsibleModule
from openpyxl import *
import xlwings as xw


def read_data(wb):
    sheetNames = wb.sheetnames
    result = {}
    for sheetName in sheetNames:
        sheet = wb[sheetName]
        ws_content = []
        for row in sheet.rows:
            ws_content.append([cell.value for cell in row])
        result[sheetName] = ws_content
    return result


def evaluate_workbook(path):
    # xlwings allows for workbook to be opened using running excel instance
    excel_app = xw.App(visible=False)
    excel_book = excel_app.books.open(path)
    excel_book.save()  # calling excel save compute functions and store it in cache
    excel_book.close()
    excel_app.quit()

    wb = load_workbook(filename=path, data_only=True)
    return read_data(wb)


def check_excel_installation():
    import platform
    # Checks if os is Windows or darwin (macOS)
    if platform.system() == 'Windows':
        try:
            # Check if Excel executable exists
            excel_path = os.path.join(os.environ["ProgramFiles"], "Microsoft Office", "root", "Office16", "EXCEL.EXE")
            excel_path_x86 = os.path.join(os.environ["ProgramFiles"], "Microsoft Office", "root", "Office16",
                                          "EXCEL.EXE")
            return os.path.exists(excel_path) or os.path.exists(excel_path_x86)
        except KeyError:
            # Environment variable not found (Excel not installed)
            return False
    elif platform.system() == 'Darwin':
        excel_path = os.path.join("/Applications", "Microsoft Excel.app")
        return os.path.exists(excel_path)
    else:
        raise RuntimeError("Only Windows and MacOS are supported for evaluation")


def main():
    module_args = dict(path=dict(type='path', required=True),
                       evaluate=dict(type='bool', default=False))
    results = {}

    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True,
    )

    if not os.path.isfile(module.params['path']):
        module.fail_json(msg="The specified excel file does not exist at " + module.params['path'])

    excel_wb = load_workbook(filename=module.params['path'], data_only=True)
    sheets_names = excel_wb.sheetnames

    if module.params['evaluate']:
        try:
            if not check_excel_installation():
                module.fail_json(msg="Excel is not installed, needed for function evaluation.")
        except RuntimeError as e:
            module.fail_json(msg=str(e))
        results['content'] = evaluate_workbook(module.params['path'])
    else:
        excel_wb = load_workbook(filename=module.params['path'], data_only=True)
        results['content'] = read_data(excel_wb)

    results['path'] = module.params['path']
    results['sheets'] = sheets_names

    module.exit_json(**results)


if __name__ == '__main__':
    main()
