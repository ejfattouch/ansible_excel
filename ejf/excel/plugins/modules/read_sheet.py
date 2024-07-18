#!/usr/bin/python
# Copyright (c) 2024 Edward-Joseph Fattouch (ejfattouch@outlook.com)
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = r'''
module: read_document
author:
    - Edward-Joseph Fattouch (@ejfattouch)
short_description: Reads data from a single sheet in an Excel document
description:
    - Reads data from a single sheet in an Excel document using the openpyxl module.
requirements:
    - "openpyxl"
    - "xlwings module and running Excel instance on windows and macOS (only required for data evaluation)"
options:
  path:
    description:
      - Path to the Excel document.
    type: str
  sheet:
    description:
      - The name of the sheet to read.
    notes: If left empty, first sheet in the workbook will be used.
    type: str
    default: ''
  evaluate:
    description:
      - Whether or not to evaluate the functions in an Excel document.
    notes: Only compatible on Windows and MacOS with xlwings and an installed Excel instance
    type: bool
    default: false
'''

EXAMPLES = r"""
- name: Read an entire Excel document
  ejf.read_excel_document:
    path: /your/path/excel/document.xlsx
  register: document
  
- name: Read an Excel document with its values evaluated
  ejf.read_excel_document:
    path: /your/path/excel/document.xlsx
    evaluate: true
  register: document
"""

RETURN = r"""
content:
    description: The contents of the sheet.
    type: list
    sample: ["Content 1", "Content 2", 1234, ...]
    returned: always
evaluated:
    description: Returns True if the functions were evaluated.
    type: bool
    returned: always
path:
    description: The path to the Excel document.
    type: str
    returned: always
sheet:
    description: Name of the sheet that was read.
    type: str
    returned: always
"""

import os
from ansible.module_utils.basic import AnsibleModule
from openpyxl import *


def read_data(wb, sheet_name):
    result = []
    sheet = wb[sheet_name]
    for row in sheet.rows:
        result.append([cell.value for cell in row])
    return result


def evaluate_workbook(path, sheet_name):
    import xlwings as xw
    # xlwings allows for workbook to be opened using running excel instance
    excel_app = xw.App(visible=False)
    excel_book = excel_app.books.open(path)
    excel_book.save()  # calling excel save compute functions and store it in cache
    excel_book.close()
    excel_app.quit()

    wb = load_workbook(filename=path, data_only=True)
    return read_data(wb, sheet_name)


def check_excel_installation():
    import platform
    # Checks if os is Windows or darwin (macOS)
    if platform.system() == 'Windows':
        try:
            # Check if Excel executable exists
            excel_path = os.path.join(os.environ["ProgramFiles"], "Microsoft Office", "root", "Office16", "EXCEL.EXE")
            excel_path_x86 = os.path.join(os.environ["ProgramFiles(x86)"], "Microsoft Office", "root", "Office16",
                                          "EXCEL.EXE")
            return os.path.exists(excel_path) or os.path.exists(excel_path_x86)
        except KeyError:  # Environment variable not found (Excel not installed)
            return False
    elif platform.system() == 'Darwin':
        excel_path = os.path.join("/Applications", "Microsoft Excel.app")
        return os.path.exists(excel_path)
    else:
        raise RuntimeError("Only Windows and MacOS are supported for evaluation")


def main():
    module_args = dict(path=dict(type='path', required=True),
                       sheet=dict(type='str', required=False, default=''),
                       evaluate=dict(type='bool', default=False))
    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=True,
    )

    if not os.path.isfile(module.params['path']):
        module.fail_json(msg="The specified excel file does not exist at " + module.params['path'])
        return 1

    results = {}
    results['path'] = os.path.abspath(module.params['path'])
    sheet_name = module.params['sheet']

    if not sheet_name:
        wb = load_workbook(filename=module.params['path'], data_only=True)
        sheet_name = wb.sheetnames[0]  # Get the first sheet if the sheet_name is unspecified

    if module.params['evaluate']:
        try:
            if not check_excel_installation():
                module.fail_json(msg="Excel is not installed, needed for function evaluation.")
                return 1
            results['content'] = evaluate_workbook(module.params['path'], sheet_name)
            results['evaluated'] = True
        except RuntimeError as e:  # Exception when wrong os is trying to be used for func validation
            module.fail_json(msg=str(e))
            return 1
        except ModuleNotFoundError as e:
            module.fail_json(msg=f"{e.name} is not installed, needed for function evaluation.")
            return 1
    else:
        try:
            excel_wb = load_workbook(filename=module.params['path'], data_only=True)
            results['content'] = read_data(excel_wb, sheet_name)
            results['evaluated'] = False
        except KeyError as e:
            err_msg = f"Worksheet '{module.params['sheet']}' does not exist in Excel workbook {os.path.abspath(module.params['path'])}"
            module.fail_json(msg=err_msg)
            return 1

    results['sheet'] = sheet_name
    module.exit_json(**results)
    return 0


if __name__ == '__main__':
    main()
